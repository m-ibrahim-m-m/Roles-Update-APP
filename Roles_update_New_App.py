import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import time
import hashlib
import json
import logging
from datetime import datetime
import math

# Configure the page
st.set_page_config(
    page_title="Multi-Excel Role Consolidator", 
    layout="wide",
    initial_sidebar_state="collapsed",
    page_icon="📊"
)

# Custom CSS for modern styling
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Global Styles */
    .stApp {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }
    
    /* Main Container */
    .main-container {
        background: rgba(255, 255, 255, 0.95);
        border-radius: 20px;
        box-shadow: 0 20px 40px rgba(0,0,0,0.1);
        backdrop-filter: blur(10px);
        margin: 20px auto;
        padding: 0;
        overflow: hidden;
    }
    
    /* Header Styles */
    .header-section {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 40px 30px;
        text-align: center;
        margin: -1rem -1rem 2rem -1rem;
    }
    
    .header-title {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0 0 10px 0;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    
    .header-subtitle {
        font-size: 1.1rem;
        opacity: 0.9;
        margin: 0;
        font-weight: 400;
    }
    
    /* Mode Selection */
    .mode-selector {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        padding: 30px;
        margin: -1rem -1rem 2rem -1rem;
        text-align: center;
    }
    
    /* Section Cards */
    .section-card {
        background: white;
        border-radius: 16px;
        padding: 30px;
        margin: 20px 0;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        border: 1px solid #e1e8ed;
    }
    
    .section-card h3 {
        color: #333;
        font-weight: 600;
        margin-bottom: 20px;
        font-size: 1.3rem;
    }
    
    /* Upload Area */
    .upload-area {
        border: 3px dashed #ddd;
        border-radius: 20px;
        padding: 40px 20px;
        text-align: center;
        background: linear-gradient(135deg, #fafbfc 0%, #f4f6f8 100%);
        margin: 20px 0;
    }
    
    /* File Item */
    .file-item {
        background: white;
        border: 1px solid #e1e8ed;
        border-radius: 12px;
        padding: 15px 20px;
        margin: 10px 0;
        display: flex;
        align-items: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    
    .file-icon {
        width: 40px;
        height: 40px;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
        color: white;
        font-size: 1.2rem;
        margin-right: 15px;
    }
    
    /* Filter Cards */
    .filter-card {
        background: white;
        border-radius: 12px;
        padding: 20px;
        margin: 15px 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        border: 1px solid #e1e8ed;
    }
    
    .filter-title {
        font-weight: 600;
        color: #333;
        margin-bottom: 12px;
        font-size: 1rem;
    }
    
    /* Status Badges */
    .status-badge {
        padding: 6px 12px;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
        display: inline-block;
    }
    
    .status-success {
        background: rgba(40, 167, 69, 0.1);
        color: #28a745;
    }
    
    .status-warning {
        background: rgba(255, 193, 7, 0.1);
        color: #ffc107;
    }
    
    /* Results Table */
    .results-container {
        background: white;
        border-radius: 16px;
        overflow: hidden;
        box-shadow: 0 4px 12px rgba(0,0,0,0.05);
        margin: 20px 0;
    }
    
    .table-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 20px;
        font-weight: 600;
        font-size: 1.1rem;
    }
    
    /* Sheet Preview Styles */
    .sheet-preview {
        background: #f8f9fa;
        border-radius: 8px;
        padding: 15px;
        margin: 10px 0;
        border-left: 4px solid #667eea;
    }
    
    .sheet-stats {
        display: flex;
        gap: 20px;
        margin: 10px 0;
        font-size: 0.9rem;
        color: #666;
    }
    
    .stat-item {
        background: rgba(102, 126, 234, 0.1);
        padding: 5px 10px;
        border-radius: 12px;
        font-weight: 500;
    }
    
    /* Filter Section Headers */
    .filter-section-header {
        background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
        padding: 15px 20px;
        border-radius: 12px;
        margin: 20px 0 15px 0;
        border-left: 4px solid #667eea;
    }
    
    .filter-section-header h4 {
        margin: 0;
        color: #333;
        font-weight: 600;
    }
    
    .filter-section-header p {
        margin: 5px 0 0 0;
        color: #666;
        font-size: 0.9rem;
    }
    
    /* Expandable sheet sections */
    .stExpander > div > div > div > div {
        padding-top: 1rem;
    }
    
    /* NO Filter Special Styling */
    .no-filter-container {
        background: linear-gradient(135deg, rgba(0, 123, 255, 0.05) 0%, rgba(108, 117, 125, 0.05) 100%);
        border: 1px solid rgba(0, 123, 255, 0.2);
        border-radius: 8px;
        padding: 10px;
        margin: 5px 0;
    }
    
    .no-filter-label {
        color: #007bff;
        font-weight: 600;
        font-size: 0.9rem;
        margin-bottom: 5px;
    }
    
    .number-count-badge {
        background: rgba(0, 123, 255, 0.1);
        color: #007bff;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: 500;
        margin-left: 8px;
    }
    
    .spinner {
        width: 50px;
        height: 50px;
        border: 4px solid #f3f3f3;
        border-top: 4px solid #667eea;
        border-radius: 50%;
        animation: spin 1s linear infinite;
        margin: 0 auto 20px;
    }
    
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    /* Button Overrides */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 12px 24px;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(102, 126, 234, 0.3);
    }
    
    .stDownloadButton > button {
        background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 12px 24px;
        font-weight: 600;
    }
    
    .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(40, 167, 69, 0.3);
    }
    
    /* Selectbox Styling */
    .stSelectbox > div > div {
        background: white;
        border: 2px solid #e1e8ed;
        border-radius: 12px;
    }
    
    /* Text Input Styling */
    .stTextInput > div > div > input {
        background: white;
        border: 2px solid #e1e8ed;
        border-radius: 12px;
        padding: 12px 16px;
        color: #007bff !important;
    }
    
    .stTextInput > div > div > input::placeholder {
        color: #6c757d !important;
    }
    
    /* Multiselect Styling */
    .stMultiSelect > div > div {
        background: white;
        border: 2px solid #e1e8ed;
        border-radius: 12px;
    }
    
    /* Action Selectbox Styling - Dark Text */
    .action-selectbox .stSelectbox > div > div > div {
        color: #000000 !important;
        font-weight: 600;
    }
    
    .action-selectbox [data-baseweb="select"] > div:first-child {
        color: #000000 !important;
    }
    
    .action-selectbox [data-baseweb="select"] input {
        color: #000000 !important;
    }
    
    .action-selectbox [data-baseweb="select"] [class*="ValueContainer"] {
        color: #000000 !important;
    }
    
    .action-selectbox [data-baseweb="select"] [class*="singleValue"] {
        color: #000000 !important;
    }
    
    .action-selectbox * {
        color: #000000 !important;
    }
    
    div[data-testid="stSelectbox"] > div > div > div {
        color: #000000 !important;
    }
    
    /* Hide Streamlit Branding */
    #MainMenu {visibility: hidden;}
    .stDeployButton {display:none;}
    footer {visibility: hidden;}
    .stApp > header {visibility: hidden;}
    
    /* Custom spacing */
    .block-container {
        padding-top: 1rem;
        max-width: 1200px;
    }
    
    /* Quality metrics cards */
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        margin: 10px 0;
    }
    
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        margin: 10px 0;
    }
    
    .metric-label {
        font-size: 0.9rem;
        opacity: 0.9;
    }
</style>
""", unsafe_allow_html=True)

# Setup logging
def setup_logging():
    """Setup application logging"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('app.log'),
            logging.StreamHandler()
        ]
    )

def log_processing_action(user_id, action, files_processed, records_processed):
    """Log processing activities"""
    logging.info(f"User {user_id} {action}. Files: {files_processed}, Records: {records_processed}")

# Initialize logging
setup_logging()

# Helper Functions
def format_file_size(size_bytes):
    """Format file size in human readable format"""
    if size_bytes == 0:
        return "0 B"
    size_names = ["B", "KB", "MB", "GB"]
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_names[i]}"

def validate_excel_file(file):
    """Validate Excel file structure and content"""
    try:
        xls = pd.ExcelFile(file)
        if not xls.sheet_names:
            raise ValueError("File contains no sheets")
        return True, xls
    except Exception as e:
        logging.error(f"Invalid Excel file {file.name}: {str(e)}")
        return False, None

@st.cache_data
def load_excel_file(file, sheet_name):
    """Cache Excel file loading for better performance"""
    return pd.read_excel(file, sheet_name=sheet_name)

def smart_sort_values(values, column_name):
    """Smart sorting for different data types, especially for NO column"""
    if not values:
        return []
    
    # Special handling for NO column - try to sort numerically
    if column_name.upper() in ['NO', 'MRP NO', 'NUMBER', 'NUM']:
        try:
            # Convert to numeric, handling various formats
            numeric_values = []
            for val in values:
                val_str = str(val).strip()
                try:
                    # Try to convert to float first, then to int if it's a whole number
                    num_val = float(val_str)
                    if num_val.is_integer():
                        numeric_values.append((int(num_val), val_str))
                    else:
                        numeric_values.append((num_val, val_str))
                except ValueError:
                    # If conversion fails, use string sorting
                    numeric_values.append((float('inf'), val_str))
            
            # Sort by numeric value, then by string representation
            numeric_values.sort(key=lambda x: (x[0], x[1]))
            return [val[1] for val in numeric_values]
        except:
            # Fallback to string sorting if numeric sorting fails
            pass
    
    # Default string sorting for other columns
    return sorted(set(str(val).strip() for val in values), key=lambda x: (x.lower(), x))

def create_template_with_validation_separated(template_sheets, validation_options):
    """Create Excel template with separated dropdown lists for each sheet"""
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Hidden validation sheet
    hidden = wb.create_sheet("ValidationLists")
    hidden.sheet_state = "hidden"
    
    # Track current column for validation lists
    current_col = 1
    validation_ranges = {}

    # Create validation lists for each sheet's columns separately
    for sheet_name, columns_data in validation_options.items():
        for col_name, options in columns_data.items():
            if not options:
                continue
                
            # Sort options
            sorted_opts = sorted(set(options), key=lambda x: str(x))
            
            # Write options to hidden sheet
            for i, val in enumerate(sorted_opts, start=1):
                hidden.cell(row=i, column=current_col, value=str(val))
            
            # Store range for this validation list
            col_letter = get_column_letter(current_col)
            validation_ranges[(sheet_name, col_name)] = f"'ValidationLists'!${col_letter}$1:${col_letter}${len(sorted_opts)}"
            current_col += 1

    # Add Action validation
    action_values = ["Add", "Remove"]
    for i, val in enumerate(action_values, start=1):
        hidden.cell(row=i, column=current_col, value=val)
    action_range = f"'ValidationLists'!${get_column_letter(current_col)}$1:${get_column_letter(current_col)}${len(action_values)}"
    current_col += 1

    # Add user-facing sheets with separate validation for each sheet
    for sheet_name, template_rows in template_sheets.items():
        ws = wb.create_sheet(sheet_name)
        
        # Get headers from first row
        if template_rows:
            headers = list(template_rows[0].keys())
            ws.append(headers)
            
            # Add template rows
            for row in template_rows:
                ws.append(list(row.values()))

            # Apply validation for each column in this sheet
            for col_num, header in enumerate(headers, start=1):
                if (sheet_name, header) in validation_ranges:
                    # Sheet-specific validation
                    dv = DataValidation(
                        type="list", 
                        formula1=validation_ranges[(sheet_name, header)], 
                        allow_blank=True
                    )
                    ws.add_data_validation(dv)
                    dv.add(f"{get_column_letter(col_num)}2:{get_column_letter(col_num)}1048576")
                elif header == "Action":
                    # Action validation (same for all sheets)
                    dv = DataValidation(
                        type="list", 
                        formula1=action_range, 
                        allow_blank=True
                    )
                    ws.add_data_validation(dv)
                    dv.add(f"{get_column_letter(col_num)}2:{get_column_letter(col_num)}1048576")

    return wb

def show_data_preview(df, max_rows=5):
    """Show interactive data preview"""
    if st.checkbox("Show Detailed Preview", key=hashlib.md5(str(df.columns).encode()).hexdigest()):
        st.dataframe(df.head(max_rows), use_container_width=True)
        
        # Column statistics
        st.subheader("📊 Column Summary")
        col_stats = pd.DataFrame({
            'Data Type': df.dtypes,
            'Non-Null Count': df.count(),
            'Null Count': df.isnull().sum(),
            'Unique Values': df.nunique()
        })
        st.dataframe(col_stats, use_container_width=True)

def create_advanced_filters(df, column_name):
    """Create advanced filtering options"""
    col1, col2, col3 = st.columns(3)
    filtered_df = df.copy()
    
    with col1:
        if st.checkbox(f"🔍 Text Filter for {column_name}", key=f"text_{column_name}"):
            search_term = st.text_input(f"Search in {column_name}", key=f"search_{column_name}")
            if search_term:
                filtered_df = filtered_df[filtered_df[column_name].astype(str).str.contains(search_term, case=False, na=False)]
    
    with col2:
        if pd.api.types.is_numeric_dtype(df[column_name]):
            min_val = float(df[column_name].min())
            max_val = float(df[column_name].max())
            if min_val != max_val:
                selected_range = st.slider(
                    f"📏 Range for {column_name}",
                    min_val,
                    max_val,
                    (min_val, max_val),
                    key=f"range_{column_name}"
                )
                filtered_df = filtered_df[
                    (filtered_df[column_name] >= selected_range[0]) & 
                    (filtered_df[column_name] <= selected_range[1])
                ]
    
    return filtered_df

def clear_session_state():
    """Clear session state while preserving important configurations"""
    keys_to_preserve = ['theme', 'language_preference']
    current_state = st.session_state.copy()
    
    for key in list(st.session_state.keys()):
        if key not in keys_to_preserve:
            del st.session_state[key]
    
    return f"Session cleared. {len(current_state) - len(keys_to_preserve)} items removed."

def export_configuration(filters, selected_sheets, user_id):
    """Export current configuration for reuse"""
    config = {
        'filters': filters,
        'selected_sheets': selected_sheets,
        'user_id': user_id,
        'timestamp': datetime.now().isoformat()
    }
    
    config_json = json.dumps(config, indent=2)
    st.download_button(
        label="📋 Export Configuration",
        data=config_json,
        file_name=f"config_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        mime="application/json"
    )

def show_quality_metrics(df):
    """Display data quality metrics"""
    st.subheader("📈 Data Quality Report")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        completeness = (1 - df.isnull().sum().sum() / (df.shape[0] * df.shape[1])) * 100
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{completeness:.1f}%</div>
            <div class="metric-label">Data Completeness</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        duplicate_rows = df.duplicated().sum()
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{duplicate_rows}</div>
            <div class="metric-label">Duplicate Rows</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        unique_users = df['User_ID'].nunique()
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{unique_users}</div>
            <div class="metric-label">Unique Users</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        data_types = df.dtypes.value_counts()
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{len(data_types)}</div>
            <div class="metric-label">Data Types</div>
        </div>
        """, unsafe_allow_html=True)

def process_large_file_in_chunks(file, chunk_size=1000):
    """Process large Excel files in chunks"""
    xls = pd.ExcelFile(file)
    all_chunks = []
    
    for sheet in xls.sheet_names:
        try:
            # Read in chunks
            chunk_reader = pd.read_excel(file, sheet_name=sheet, chunksize=chunk_size)
            for chunk in chunk_reader:
                # Process chunk here (add your processing logic)
                processed_chunk = chunk  # Placeholder for actual processing
                all_chunks.append(processed_chunk)
        except Exception as e:
            logging.error(f"Error processing sheet {sheet} in chunks: {str(e)}")
            continue
    
    if all_chunks:
        return pd.concat(all_chunks, ignore_index=True)
    return pd.DataFrame()

def process_single_user_files(uploaded_files, user_id, filters, selected_sheets):
    """Process files for single user mode"""
    all_filtered = []
    
    for file in uploaded_files:
        try:
            is_valid, xls = validate_excel_file(file)
            if not is_valid:
                continue
                
            for sheet in xls.sheet_names:
                # Skip if this sheet is not selected
                sheet_key = f"{file.name} - {sheet}"
                if sheet_key not in selected_sheets:
                    continue
                    
                df = load_excel_file(file, sheet)
                
                if df.empty:
                    continue
                
                filtered = df.copy()
                selection_made = False
                
                # Apply filters - check both global filters and sheet-specific filters
                for filter_key, filter_values in filters.items():
                    if filter_values and filter_key in df.columns:
                        # Special handling for NO column to support multiple number selection
                        if filter_key.upper() in ['NO', 'MRP NO', 'NUMBER', 'NUM']:
                            # Convert both filter values and data to strings for comparison
                            filter_strings = [str(val).strip() for val in filter_values]
                            filtered = filtered[filtered[filter_key].astype(str).str.strip().isin(filter_strings)]
                        else:
                            filtered = filtered[filtered[filter_key].astype(str).isin(filter_values)]
                        selection_made = True
                
                # Apply sheet-specific filters
                sheet_filters_key = f"sheet_filters_{file.name}_{sheet}"
                if sheet_filters_key in filters:
                    sheet_specific_filters = filters[sheet_filters_key]
                    for col_name, selected_values in sheet_specific_filters.items():
                        if selected_values and col_name in df.columns:
                            # Special handling for NO column
                            if col_name.upper() in ['NO', 'MRP NO', 'NUMBER', 'NUM']:
                                filter_strings = [str(val).strip() for val in selected_values]
                                filtered = filtered[filtered[col_name].astype(str).str.strip().isin(filter_strings)]
                            else:
                                filtered = filtered[filtered[col_name].astype(str).isin(selected_values)]
                            selection_made = True
                
                if selection_made and not filtered.empty:
                    filtered.insert(0, "User_ID", user_id)
                    filtered["Source_File"] = file.name
                    filtered["Source_Sheet"] = sheet
                    action = st.session_state.get(f"{sheet_key}_action", "Add")
                    filtered["Action"] = action
                    all_filtered.append(filtered)
        
        except Exception as e:
            logging.error(f"Error processing {file.name}: {str(e)}")
            st.error(f"Error processing {file.name}: {str(e)}")
    
    return all_filtered

# Initialize session state
if 'consolidated_data' not in st.session_state:
    st.session_state.consolidated_data = None
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []

# Header Section
st.markdown("""
<div class="header-section">
    <h1 class="header-title">📊 Multi-Excel Role Consolidator</h1>
    <p class="header-subtitle">Upload, filter, and consolidate Excel files with advanced role management</p>
</div>
""", unsafe_allow_html=True)

# Mode Selection
st.markdown("""
<div class="mode-selector">
    <h3 style="margin-bottom: 15px; color: #333;">Choose Processing Mode</h3>
</div>
""", unsafe_allow_html=True)

mode = st.radio(
    "Select Mode:",
    ["👤 Single User Mode", "👥 Mass Upload Mode"],
    horizontal=True,
    label_visibility="collapsed"
)

# Mode description
if "Single User" in mode:
    st.info("🔍 **Single User Mode**: Process files for individual users with real-time filtering and immediate results.")
else:
    st.info("📋 **Mass Upload Mode**: Upload multiple user requests at once using a pre-configured template with separated dropdowns for each sheet.")

# File Upload Section
st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown("### 📁 File Upload")

uploaded_files = st.file_uploader(
    "Upload your Excel files (.xlsx format)",
    type=["xlsx"],
    accept_multiple_files=True,
    help="You can upload multiple Excel files at once. Each file will be processed according to your selected mode."
)

if uploaded_files:
    st.session_state.uploaded_files = uploaded_files
    
    # Display uploaded files with modern styling
    st.markdown("#### 📂 Uploaded Files")
    for i, file in enumerate(uploaded_files):
        col1, col2, col3 = st.columns([3, 2, 1])
        with col1:
            st.markdown(f"""
            <div class="file-item">
                <div class="file-icon">📄</div>
                <div>
                    <div style="font-weight: 600; color: #333;">{file.name}</div>
                    <div style="color: #666; font-size: 0.9rem;">{format_file_size(file.size)}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

# Initialize filters
filters = {}

# Single User Mode
if "Single User" in mode and uploaded_files:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown("### 👤 Single User Configuration")
    
    # User ID Input
    user_id = st.text_input(
        "User ID",
        placeholder="Enter unique user identifier",
        help="This will be added to all consolidated records for this user"
    )
    
    if user_id:
        # Create filter interface
        st.markdown("#### 📋 Sheet Selection")
        
        # First, let users select which sheets to process
        sheet_options = []
        file_sheet_mapping = {}
        all_sheet_data = {}
        
        for file in uploaded_files:
            try:
                is_valid, xls = validate_excel_file(file)
                if not is_valid:
                    continue
                    
                file_sheet_mapping[file.name] = []
                for sheet in xls.sheet_names:
                    sheet_key = f"{file.name} - {sheet}"
                    sheet_options.append(sheet_key)
                    file_sheet_mapping[file.name].append(sheet)
                    
                    # Store sheet data for later filtering
                    df = load_excel_file(file, sheet)
                    if not df.empty:
                        all_sheet_data[sheet_key] = df
            except Exception as e:
                st.error(f"Error analyzing {file.name}: {str(e)}")
        
        # Sheet selection multiselect
        selected_sheets = st.multiselect(
            "📄 Select Sheets to Process",
            options=sheet_options,
            default=sheet_options,  # Select all by default
            help="Choose which sheets from your uploaded files to include in the consolidation"
        )
        
        if selected_sheets:
            # Sheet-specific filters
            st.markdown("#### 🎯 Sheet-Specific Filters")
            st.markdown("*Configure individual filters for each selected sheet*")
            
            # Create expandable sections for each selected sheet
            for sheet_key in selected_sheets:
                if sheet_key in all_sheet_data:
                    df = all_sheet_data[sheet_key]
                    
                    with st.expander(f"📊 {sheet_key} ({len(df)} rows)", expanded=False):
                        # Show preview of the sheet
                        st.markdown("**Data Preview:**")
                        show_data_preview(df)
                        
                        # Create filters specific to this sheet
                        sheet_cols = st.columns(2)
                        sheet_filter_key = f"sheet_filters_{sheet_key.split(' - ')[0]}_{sheet_key.split(' - ')[1]}"
                        
                        if sheet_filter_key not in filters:
                            filters[sheet_filter_key] = {}
                        
                        # Get unique values for each column in this sheet
                        available_columns = [col for col in df.columns if col not in ['User_ID', 'Source_File', 'Source_Sheet', 'TOTAL']]
                        
                        for idx, col in enumerate(available_columns[:4]):  # Limit to first 4 columns, excluding TOTAL
                            if col.upper() in ['NO', 'MRP NO', 'NUMBER', 'NUM', 'TOTAL']:
                                continue  # Skip filtering for NO and TOTAL columns
                            with sheet_cols[idx % 2]:
                                # Get unique values for this column
                                unique_values = df[col].dropna().unique().tolist()
                                
                                if unique_values:
                                    # Use smart sorting
                                    sorted_values = smart_sort_values(unique_values, col)
                                    
                                    label = f"🔹 {col}"
                                    help_text = f"Filter {sheet_key} by {col}"
                                    
                                    filters[sheet_filter_key][col] = st.multiselect(
                                        label,
                                        options=sorted_values,
                                        key=f"{sheet_key}_{col}_filter",
                                        help=help_text
                                    )
                        
                        # Add Action selection with custom styling
                        st.markdown("#### Action Selection")
                        st.markdown('<div class="action-selectbox">', unsafe_allow_html=True)
                        st.selectbox(
                            "Choose action for this sheet:",
                            ["Add", "Remove"],
                            key=f"{sheet_key}_action",
                            help="Select action for this sheet"
                        )
                        st.markdown('</div>', unsafe_allow_html=True)
            
            # Configuration export
            st.markdown("#### ⚙️ Configuration Management")
            export_configuration(filters, selected_sheets, user_id)
            
            # Reset button
            if st.button("🔄 Reset Session"):
                clear_session_state()
                st.rerun()
        
        else:
            st.warning("⚠️ Please select at least one sheet to process.")
        
        # Process Button
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🔄 Process & Consolidate", type="primary", use_container_width=True):
                if not selected_sheets:
                    st.error("❌ Please select at least one sheet to process.")
                else:
                    # Show processing animation
                    with st.spinner("Processing your selected sheets..."):
                        progress_bar = st.progress(0)
                        
                        # Simulate progress for each selected sheet
                        total_sheets = len(selected_sheets)
                        for i in range(total_sheets):
                            time.sleep(0.1)  # Small delay for visual effect
                            progress_bar.progress((i + 1) / total_sheets)
                        
                        # Process files
                        all_filtered = process_single_user_files(uploaded_files, user_id, filters, selected_sheets)
                        
                        if all_filtered:
                            consolidated = pd.concat(all_filtered, ignore_index=True)
                            
                            # Apply business rules
                            consolidated = consolidated[consolidated["Source_File"] != "PLANT ALL.xlsx"]
                            consolidated = consolidated.dropna(axis=1, how="all")
                            
                            st.session_state.consolidated_data = consolidated
                            
                            # Log the processing action
                            log_processing_action(user_id, "processed files", len(uploaded_files), len(consolidated))
                            
                            # Show detailed success message
                            unique_users = consolidated['User_ID'].nunique()
                            total_records = len(consolidated)
                            processed_sheets = len(consolidated.groupby(['Source_File', 'Source_Sheet']))
                            
                            st.success(
                                f"✅ **Processing Complete!**\n\n"
                                f"📊 **{total_records}** records consolidated\n"
                                f"👤 **{unique_users}** unique user(s)\n"
                                f"📄 **{processed_sheets}** sheet(s) processed\n"
                                f"📁 **{len(selected_sheets)}** sheet(s) selected"
                            )
                        else:
                            st.warning("⚠️ No data matched your filter criteria. Please adjust your selections.")
                            st.info("💡 **Tips:**\n- Try selecting different filter values\n- Check if the selected sheets contain the expected data\n- Ensure at least one filter is applied")
    
    st.markdown('</div>', unsafe_allow_html=True)

# Mass Upload Mode
elif "Mass Upload" in mode and uploaded_files:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown("### 📋 Mass Upload Process")
    
    # Step 1: Template Generation
    st.markdown("#### Step 1: Download Template")
    st.markdown("Generate a template with **separated dropdown lists for each sheet** based on your uploaded files.")
    
    if st.button("📥 Generate & Download Template", type="primary"):
        with st.spinner("Generating template with separated dropdowns..."):
            template_sheets = {}
            validation_options = {}
            
            for file in uploaded_files:
                try:
                    is_valid, xls = validate_excel_file(file)
                    if not is_valid:
                        continue
                        
                    for sheet in xls.sheet_names:
                        df = load_excel_file(file, sheet)
                        if df.empty:
                            continue
                        
                        # Filterable columns, excluding NO and TOTAL
                        filter_cols = [df.columns[0]]
                        for col in ["PLANT", "APP"]:
                            if col in df.columns and col != "TOTAL":
                                filter_cols.append(col)
                        
                        # Initialize sheet in validation options
                        if sheet not in validation_options:
                            validation_options[sheet] = {}
                        
                        # Collect dropdown options for this sheet's columns
                        for col in filter_cols:
                            if col not in validation_options[sheet]:
                                validation_options[sheet][col] = set()
                            validation_options[sheet][col].update(df[col].dropna().astype(str).unique())
                        
                        # Build template row
                        row = {"User_ID": "", "Source_File": file.name, "Source_Sheet": sheet}
                        for col in filter_cols:
                            row[col] = ""
                        row["Action"] = ""
                        
                        if sheet not in template_sheets:
                            template_sheets[sheet] = []
                        template_sheets[sheet].append(row)
                
                except Exception as e:
                    logging.error(f"Error processing {file.name} for template: {str(e)}")
                    st.error(f"Error processing {file.name}: {str(e)}")
            
            if template_sheets:
                wb = create_template_with_validation_separated(template_sheets, validation_options)
                buffer = BytesIO()
                wb.save(buffer)
                
                st.download_button(
                    label="📥 Download Mass Upload Template",
                    data=buffer.getvalue(),
                    file_name="Mass_Upload_Template_Separated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success("✅ Template generated successfully with separated dropdowns for each sheet!")
                st.info("💡 **Template Features:**\n- Separate dropdown lists for each sheet\n- Data validation for accurate input\n- Action column with Add/Remove options")
            else:
                st.error("❌ No valid data found in uploaded files to generate template.")
    
    # Step 2: Upload Filled Template
    st.markdown("#### Step 2: Upload Filled Template")
    st.markdown("Upload your completed template file to process multiple users at once.")
    
    mass_file = st.file_uploader(
        "Upload filled template",
        type=["xlsx"],
        key="mass_upload",
        help="Upload the template file after filling it with user requirements"
    )
    
    if mass_file:
        if st.button("🔄 Process Mass Upload", type="primary"):
            with st.spinner("Processing mass upload..."):
                progress_bar = st.progress(0)
                all_filtered = []
                
                try:
                    is_valid, xls = validate_excel_file(mass_file)
                    if not is_valid:
                        st.error("❌ Invalid template file uploaded.")
                    else:
                        total_sheets = len(xls.sheet_names)
                        
                        for sheet_idx, sheet in enumerate(xls.sheet_names):
                            df_template = pd.read_excel(mass_file, sheet_name=sheet)
                            
                            if "User_ID" not in df_template.columns:
                                continue
                            
                            for _, row in df_template.iterrows():
                                if pd.isna(row["User_ID"]) or str(row["User_ID"]).strip() == "":
                                    continue
                                
                                # Find corresponding source file
                                file_name = row["Source_File"]
                                sheet_name = row["Source_Sheet"]
                                
                                source_file = None
                                for f in uploaded_files:
                                    if f.name == file_name:
                                        source_file = f
                                        break
                                
                                if source_file:
                                    df = load_excel_file(source_file, sheet_name)
                                    if df.empty:
                                        continue
                                    
                                    filtered = df.copy()
                                    selection_made = False
                                    
                                    # Apply filters from template, excluding NO, TOTAL and Action
                                    for col in df.columns:
                                        if col not in ['TOTAL', 'Action'] and col in row and pd.notna(row[col]) and str(row[col]).strip() != "":
                                            # Special handling for NO columns
                                            if col.upper() in ['NO', 'MRP NO', 'NUMBER', 'NUM']:
                                                filtered = filtered[filtered[col].astype(str).str.strip() == str(row[col]).strip()]
                                            else:
                                                filtered = filtered[filtered[col].astype(str) == str(row[col])]
                                            selection_made = True
                                    
                                    if selection_made and not filtered.empty:
                                        filtered.insert(0, "User_ID", row["User_ID"])
                                        filtered["Source_File"] = file_name
                                        filtered["Source_Sheet"] = sheet_name
                                        if "Action" in row and pd.notna(row["Action"]):
                                            filtered["Action"] = row["Action"]
                                        all_filtered.append(filtered)
                            
                            progress_bar.progress((sheet_idx + 1) / total_sheets)
                        
                        if all_filtered:
                            consolidated = pd.concat(all_filtered, ignore_index=True)
                            consolidated = consolidated[consolidated["Source_File"] != "PLANT ALL.xlsx"]
                            consolidated = consolidated.dropna(axis=1, how="all")
                            
                            st.session_state.consolidated_data = consolidated
                            
                            # Log the mass upload processing
                            log_processing_action("MASS_UPLOAD", "processed mass upload", len(uploaded_files), len(consolidated))
                            
                            st.success(f"✅ Successfully processed {len(consolidated)} records for {len(consolidated['User_ID'].unique())} users!")
                        else:
                            st.warning("⚠️ No valid data found in the template file.")
                
                except Exception as e:
                    logging.error(f"Error processing mass upload: {str(e)}")
                    st.error(f"Error processing mass upload: {str(e)}")
    
    st.markdown('</div>', unsafe_allow_html=True)

# Results Section
if st.session_state.consolidated_data is not None:
    st.markdown('<div class="results-container">', unsafe_allow_html=True)
    st.markdown('<div class="table-header">📊 Consolidated Results</div>', unsafe_allow_html=True)
    
    df = st.session_state.consolidated_data
    
    # Display quality metrics
    show_quality_metrics(df)
    
    # Display summary metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Records", len(df))
    with col2:
        st.metric("Unique Users", df['User_ID'].nunique())
    with col3:
        st.metric("Source Files", df['Source_File'].nunique())
    with col4:
        st.metric("Sheets Processed", df['Source_Sheet'].nunique())
    
    # Display the consolidated data
    st.dataframe(df, use_container_width=True, height=400)
    
    # Download consolidated results
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Consolidated")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.download_button(
            label="📥 Download Consolidated Excel",
            data=output.getvalue(),
            file_name="consolidated_roles.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    
    st.markdown('</div>', unsafe_allow_html=True)

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #444; padding: 20px;'>"
    "Built with ❤️ using Streamlit | Multi-Excel Role Consolidator v3.0"
    "</div>",
    unsafe_allow_html=True
)
