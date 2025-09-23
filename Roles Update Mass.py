if st.button("📥 Generate & Download Template", type="primary"):                            else:
                                st.warning(f"⚠️ Source file '{file_name}' not found in uploaded files.")import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import time
import re

# Configure the page
st.set_page_config(
    page_title="Excel Role Consolidator", 
    layout="wide",
    initial_sidebar_state="collapsed",
    page_icon="📊"
)

# Custom CSS for modern styling
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Global Styles */
    .stApp {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }
    
    /* Main Container */
    .main-container {
        background: rgba(255, 255, 255, 0.95);
        border-radius: 20px;
        box-shadow: 0 20px 40px rgba(0,0,0,0.1);
        backdrop-filter: blur(10px);
        margin: 20px auto;
        padding: 0;
        overflow: hidden;
    }
    
    /* Header Styles */
    .header-section {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 40px 30px;
        text-align: center;
        margin: -1rem -1rem 2rem -1rem;
    }
    
    .header-title {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0 0 10px 0;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    
    .header-subtitle {
        font-size: 1.1rem;
        opacity: 0.9;
        margin: 0;
        font-weight: 400;
    }
    
    /* Section Cards */
    .section-card {
        background: white;
        border-radius: 16px;
        padding: 30px;
        margin: 20px 0;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        border: 1px solid #e1e8ed;
    }
    
    .section-card h3 {
        color: #333;
        font-weight: 600;
        margin-bottom: 20px;
        font-size: 1.3rem;
    }
    
    /* Upload Area */
    .upload-area {
        border: 3px dashed #ddd;
        border-radius: 20px;
        padding: 40px 20px;
        text-align: center;
        background: linear-gradient(135deg, #fafbfc 0%, #f4f6f8 100%);
        margin: 20px 0;
    }
    
    /* File Item */
    .file-item {
        background: white;
        border: 1px solid #e1e8ed;
        border-radius: 12px;
        padding: 15px 20px;
        margin: 10px 0;
        display: flex;
        align-items: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    
    .file-icon {
        width: 40px;
        height: 40px;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
        color: white;
        font-size: 1.2rem;
        margin-right: 15px;
    }
    
    /* Filter Cards */
    .filter-card {
        background: white;
        border-radius: 12px;
        padding: 20px;
        margin: 15px 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        border: 1px solid #e1e8ed;
    }
    
    .filter-title {
        font-weight: 600;
        color: #333;
        margin-bottom: 12px;
        font-size: 1rem;
    }
    
    /* Status Badges */
    .status-badge {
        padding: 6px 12px;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
        display: inline-block;
    }
    
    .status-success {
        background: rgba(40, 167, 69, 0.1);
        color: #28a745;
    }
    
    .status-warning {
        background: rgba(255, 193, 7, 0.1);
        color: #ffc107;
    }
    
    /* Results Table */
    .results-container {
        background: white;
        border-radius: 16px;
        overflow: hidden;
        box-shadow: 0 4px 12px rgba(0,0,0,0.05);
        margin: 20px 0;
    }
    
    .table-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 20px;
        font-weight: 600;
        font-size: 1.1rem;
    }
    
    /* Sheet Preview Styles */
    .sheet-preview {
        background: #f8f9fa;
        border-radius: 8px;
        padding: 15px;
        margin: 10px 0;
        border-left: 4px solid #667eea;
    }
    
    .sheet-stats {
        display: flex;
        gap: 20px;
        margin: 10px 0;
        font-size: 0.9rem;
        color: #666;
    }
    
    .stat-item {
        background: rgba(102, 126, 234, 0.1);
        padding: 5px 10px;
        border-radius: 12px;
        font-weight: 500;
    }
    
    /* Filter Section Headers */
    .filter-section-header {
        background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
        padding: 15px 20px;
        border-radius: 12px;
        margin: 20px 0 15px 0;
        border-left: 4px solid #667eea;
    }
    
    .filter-section-header h4 {
        margin: 0;
        color: #333;
        font-weight: 600;
    }
    
    .filter-section-header p {
        margin: 5px 0 0 0;
        color: #666;
        font-size: 0.9rem;
    }
    
    /* Expandable sheet sections */
    .stExpander > div > div > div > div {
        padding-top: 1rem;
    }
    
    /* NO Filter Special Styling */
    .no-filter-container {
        background: linear-gradient(135deg, rgba(0, 123, 255, 0.05) 0%, rgba(108, 117, 125, 0.05) 100%);
        border: 1px solid rgba(0, 123, 255, 0.2);
        border-radius: 8px;
        padding: 10px;
        margin: 5px 0;
    }
    
    .no-filter-label {
        color: #007bff;
        font-weight: 600;
        font-size: 0.9rem;
        margin-bottom: 5px;
    }
    
    .number-count-badge {
        background: rgba(0, 123, 255, 0.1);
        color: #007bff;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: 500;
        margin-left: 8px;
    }
    
    .spinner {
        width: 50px;
        height: 50px;
        border: 4px solid #f3f3f3;
        border-top: 4px solid #667eea;
        border-radius: 50%;
        animation: spin 1s linear infinite;
        margin: 0 auto 20px;
    }
    
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    /* Button Overrides */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 12px 24px;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(102, 126, 234, 0.3);
    }
    
    .stDownloadButton > button {
        background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 12px 24px;
        font-weight: 600;
    }
    
    .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(40, 167, 69, 0.3);
    }
    
    /* Toggle Styling */
    .stToggle > div > div > div > div {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    
    /* Expander Styling */
    .stExpander {
        border: 1px solid #e1e8ed;
        border-radius: 12px;
        overflow: hidden;
    }
    
    .stExpander > div:first-child {
        background: linear-gradient(135deg, rgba(102, 126, 234, 0.05) 0%, rgba(118, 75, 162, 0.05) 100%);
        padding: 12px 20px;
        border-bottom: 1px solid #e1e8ed;
    }
    
    .stExpander > div:first-child > div > p {
        font-weight: 600;
        color: #333;
        margin: 0;
    }
    
    /* File Count Badge */
    .file-count-badge {
        background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
        color: #667eea;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 0.9rem;
        font-weight: 600;
        border: 1px solid rgba(102, 126, 234, 0.2);
    }
    .stSelectbox > div > div {
        background: white;
        border: 2px solid #e1e8ed;
        border-radius: 12px;
    }
    
    /* Text Input Styling */
    .stTextInput > div > div > input {
        background: white;
        border: 2px solid #e1e8ed;
        border-radius: 12px;
        padding: 12px 16px;
    }
    
    /* Multiselect Styling */
    .stMultiSelect > div > div {
        background: white;
        border: 2px solid #e1e8ed;
        border-radius: 12px;
    }
    
    /* Hide Streamlit Branding */
    #MainMenu {visibility: hidden;}
    .stDeployButton {display:none;}
    footer {visibility: hidden;}
    .stApp > header {visibility: hidden;}
    
    /* Custom spacing */
    .block-container {
        padding-top: 1rem;
        max-width: 1200px;
    }
</style>
""", unsafe_allow_html=True)

# Helper Functions
def format_file_size(size_bytes):
    """Format file size in human readable format"""
    if size_bytes == 0:
        return "0 B"
    size_names = ["B", "KB", "MB", "GB"]
    import math
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_names[i]}"

def smart_sort_values(values, column_name):
    """Smart sorting for different data types, especially for NO column"""
    if not values:
        return []
    
    # Special handling for NO column - try to sort numerically
    if column_name.upper() in ['NO', 'MRP NO', 'NUMBER', 'NUM']:
        try:
            # Convert to numeric, handling various formats
            numeric_values = []
            for val in values:
                val_str = str(val).strip()
                try:
                    # Try to convert to float first, then to int if it's a whole number
                    num_val = float(val_str)
                    if num_val.is_integer():
                        numeric_values.append((int(num_val), val_str))
                    else:
                        numeric_values.append((num_val, val_str))
                except ValueError:
                    # If conversion fails, use string sorting
                    numeric_values.append((float('inf'), val_str))
            
            # Sort by numeric value, then by string representation
            numeric_values.sort(key=lambda x: (x[0], x[1]))
            return [val[1] for val in numeric_values]
        except:
            # Fallback to string sorting if numeric sorting fails
            pass
    
    # Default string sorting for other columns
    return sorted(set(str(val).strip() for val in values), key=lambda x: (x.lower(), x))

def create_sheet_specific_template(template_sheets, validation_options):
    """Create Excel template with sheet-specific data validation"""
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Create a hidden validation sheet for each source sheet
    validation_sheets = {}
    
    # First, create all the validation sheets
    for sheet_key, options in validation_options.items():
        # Create a safe name for the validation sheet
        safe_name = re.sub(r'[\\/*?[\]:]', '', f"Validation_{sheet_key}")[:30]
        hidden = wb.create_sheet(safe_name)
        hidden.sheet_state = "hidden"
        validation_sheets[sheet_key] = {"sheet": hidden, "ranges": {}}
        
        # Add validation data for each column
        col_index = 1
        for col_name, col_options in options.items():
            if not col_options:
                continue
                
            sorted_opts = sorted(set(col_options), key=lambda x: str(x))
            for i, val in enumerate(sorted_opts, start=1):
                hidden.cell(row=i, column=col_index, value=str(val))
            
            col_letter = chr(64 + col_index)
            validation_sheets[sheet_key]["ranges"][col_name] = f"'{safe_name}'!${col_letter}$1:${col_letter}${len(sorted_opts)}"
            col_index += 1
        
        # Add Actions validation (Add/Remove) for all sheets
        actions_data = ["Add", "Remove"]
        for i, val in enumerate(actions_data, start=1):
            hidden.cell(row=i, column=col_index, value=val)
        
        col_letter = chr(64 + col_index)
        validation_sheets[sheet_key]["ranges"]["Actions"] = f"'{safe_name}'!${col_letter}$1:${col_letter}$2"

    # Add user-facing sheets with data validation
    for sheet_key, template_rows in template_sheets.items():
        # Create a safe name for the user sheet
        safe_sheet_name = re.sub(r'[\\/*?[\]:]', '', sheet_key)[:30]
        ws = wb.create_sheet(safe_sheet_name)
        
        headers = list(template_rows[0].keys())
        ws.append(headers)
        
        for row in template_rows:
            ws.append(list(row.values()))

        # Apply validation from the corresponding validation sheet
        if sheet_key in validation_sheets:
            for col_num, header in enumerate(headers, start=1):
                if header in validation_sheets[sheet_key]["ranges"]:
                    dv = DataValidation(
                        type="list", 
                        formula1=validation_sheets[sheet_key]["ranges"][header], 
                        allow_blank=True
                    )
                    ws.add_data_validation(dv)
                    dv.add(f"{chr(64 + col_num)}2:{chr(64 + col_num)}1048576")

    return wb

# Initialize session state
if 'consolidated_data' not in st.session_state:
    st.session_state.consolidated_data = None
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []

# Header Section
st.markdown("""
<div class="header-section">
    <h1 class="header-title">📊 Excel Role Consolidator</h1>
    <p class="header-subtitle">Upload Excel files and consolidate roles using a template-based approach</p>
</div>
""", unsafe_allow_html=True)

# File Upload Section
st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown("### 📁 File Upload")

uploaded_files = st.file_uploader(
    "Upload your Excel files (.xlsx format)",
    type=["xlsx"],
    accept_multiple_files=True,
    help="Upload the Excel files you want to use as data sources for role consolidation."
)

if uploaded_files:
    st.session_state.uploaded_files = uploaded_files
    
    # Add toggle for showing/hiding uploaded files
    col1, col2 = st.columns([1, 4])
    with col1:
        show_files = st.toggle("Show Files", value=True, key="show_uploaded_files")
    with col2:
        st.markdown(f'<div class="file-count-badge">{len(uploaded_files)} file(s) uploaded</div>', unsafe_allow_html=True)
    
    # Display uploaded files with modern styling (only if toggled on)
    if show_files:
        st.markdown("#### 📂 Uploaded Files")
        
        # Create an expander for better organization
        with st.expander("📁 File Details", expanded=True):
            for i, file in enumerate(uploaded_files):
                col1, col2, col3 = st.columns([3, 2, 1])
                with col1:
                    st.markdown(f"""
                    <div class="file-item">
                        <div class="file-icon">📄</div>
                        <div>
                            <div style="font-weight: 600; color: #333;">{file.name}</div>
                            <div style="color: #666; font-size: 0.9rem;">{format_file_size(file.size)}</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

# Mass Upload Mode
if uploaded_files:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown("### 📋 Template-Based Role Consolidation")
    
    # Step 1: Template Generation
    st.markdown("#### Step 1: Download Template")
    st.markdown("Generate a template with validation dropdowns based on your uploaded files. Each sheet will have its own dropdown lists.")
    
    # Add a debug section to show sheet name mapping
    if st.button("🔍 Debug: Show Sheet Name Mapping", help="Show how template sheet names map to source files"):
        if uploaded_files:
            st.markdown("#### 🗂️ Sheet Name Mapping")
            with st.expander("Sheet Name Details", expanded=True):
                mapping_data = []
                for file in uploaded_files:
                    try:
                        xls = pd.ExcelFile(file)
                        for sheet in xls.sheet_names:
                            template_name = f"{file.name} - {sheet}".strip()
                            mapping_data.append({
                                "Template Sheet Name": template_name,
                                "Source File": file.name,
                                "Source Sheet": sheet,
                                "Template Name Length": len(template_name)
                            })
                    except Exception as e:
                        st.error(f"Error reading {file.name}: {str(e)}")
                
                if mapping_data:
                    mapping_df = pd.DataFrame(mapping_data)
                    st.dataframe(mapping_df, use_container_width=True)
                    
                    # Show warning for long names
                    long_names = mapping_df[mapping_df["Template Name Length"] > 31]
                    if not long_names.empty:
                        st.warning("⚠️ Some template sheet names are longer than 31 characters and may be truncated by Excel:")
                        st.dataframe(long_names[["Template Sheet Name", "Template Name Length"]], use_container_width=True)
        with st.spinner("Generating template with sheet-specific dropdowns..."):
            template_sheets = {}
            validation_options = {}
            
            for file in uploaded_files:
                try:
                    xls = pd.ExcelFile(file)
                    for sheet in xls.sheet_names:
                        df = pd.read_excel(file, sheet_name=sheet)
                        if df.empty:
                            continue
                        
                        # Create a unique key for this sheet (clean up the naming)
                        sheet_key = f"{file.name} - {sheet}".strip()
                        
                        # Filterable columns (excluding Total column)
                        filter_cols = [df.columns[0]]
                        for col in ["PLANT", "APP", "NO"]:
                            if col in df.columns:
                                filter_cols.append(col)
                        
                        # Collect dropdown options for this specific sheet
                        if sheet_key not in validation_options:
                            validation_options[sheet_key] = {}
                            
                        for col in filter_cols:
                            if col not in validation_options[sheet_key]:
                                validation_options[sheet_key][col] = set()
                            
                            # Special handling for NO column
                            if col == "NO":
                                values = df[col].dropna().astype(str).str.strip()
                                validation_options[sheet_key][col].update(values)
                            else:
                                validation_options[sheet_key][col].update(df[col].dropna().astype(str))
                        
                        # Build template row (without Total, Source_File, Source_Sheet columns, but with Actions column)
                        row = {"User_ID": "", "Actions": ""}
                        for col in filter_cols:
                            row[col] = ""
                        
                        if sheet_key not in template_sheets:
                            template_sheets[sheet_key] = []
                        template_sheets[sheet_key].append(row)
                
                except Exception as e:
                    st.error(f"Error processing {file.name}: {str(e)}")
            
            if template_sheets:
                wb = create_sheet_specific_template(template_sheets, validation_options)
                buffer = BytesIO()
                wb.save(buffer)
                
                st.download_button(
                    label="📥 Download Template with Sheet-Specific Dropdowns",
                    data=buffer.getvalue(),
                    file_name="Sheet_Specific_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success("✅ Template generated successfully! Each sheet has its own dropdown lists based on the source data.")
                st.info("💡 **Note:** The template contains separate validation for each source sheet, with an 'Actions' column (Add/Remove) and excludes the 'Total', 'Source_File', and 'Source_Sheet' columns. File and sheet information is embedded in the sheet names.")
            else:
                st.error("❌ Could not generate template. Please check if your Excel files contain valid data.")
    
    # Step 2: Upload Filled Template
    st.markdown("#### Step 2: Upload Filled Template")
    st.markdown("Upload your completed template file to process multiple users at once.")
    
    mass_file = st.file_uploader(
        "Upload filled template",
        type=["xlsx"],
        key="mass_upload",
        help="Upload the template file after filling it with user requirements"
    )
    
    if mass_file:
        if st.button("🔄 Process Mass Upload", type="primary"):
            with st.spinner("Processing mass upload..."):
                progress_bar = st.progress(0)
                all_filtered = []
                
                try:
                    xls = pd.ExcelFile(mass_file)
                    total_sheets = len(xls.sheet_names)
                    
                    for sheet_idx, sheet in enumerate(xls.sheet_names):
                        # Skip validation sheets
                        if sheet.startswith("Validation_"):
                            continue
                            
                        df_template = pd.read_excel(mass_file, sheet_name=sheet)
                        
                        if "User_ID" not in df_template.columns:
                            continue
                        
                        for _, row in df_template.iterrows():
                            if pd.isna(row["User_ID"]) or str(row["User_ID"]).strip() == "":
                                continue
                            
                            # Extract sheet information from sheet name (format: "filename - sheetname")
                            # Handle long sheet names by cleaning up spaces
                            sheet_parts = sheet.split(" - ")
                            if len(sheet_parts) >= 2:
                                file_name = sheet_parts[0].strip()
                                sheet_name = " - ".join(sheet_parts[1:]).strip()  # Handle cases where sheet name contains " - "
                            else:
                                # Fallback: try to match with uploaded files
                                file_name = None
                                sheet_name = sheet.strip()
                                for f in uploaded_files:
                                    file_base_name = f.name.replace(".xlsx", "").strip()
                                    if file_base_name in sheet:
                                        file_name = f.name
                                        # Remove file name from sheet name and clean up
                                        sheet_name = sheet.replace(file_base_name, "").strip()
                                        # Remove leading/trailing dashes and spaces
                                        sheet_name = sheet_name.strip(" -").strip()
                                        break
                            
                            source_file = None
                            for f in uploaded_files:
                                if f.name == file_name:
                                    source_file = f
                                    break
                            
                            if source_file:
                                try:
                                    # First, check if the sheet exists in the source file
                                    source_xls = pd.ExcelFile(source_file)
                                    available_sheets = source_xls.sheet_names
                                    
                                    # Try exact match first
                                    if sheet_name in available_sheets:
                                        df = pd.read_excel(source_file, sheet_name=sheet_name)
                                    else:
                                        # Try fuzzy matching for sheet names
                                        matching_sheet = None
                                        for available_sheet in available_sheets:
                                            # Check if sheet_name is contained in available_sheet or vice versa
                                            if (sheet_name.lower() in available_sheet.lower() or 
                                                available_sheet.lower() in sheet_name.lower()):
                                                matching_sheet = available_sheet
                                                break
                                        
                                        if matching_sheet:
                                            df = pd.read_excel(source_file, sheet_name=matching_sheet)
                                            st.warning(f"⚠️ Sheet '{sheet_name}' not found in {file_name}. Using '{matching_sheet}' instead.")
                                        else:
                                            st.error(f"❌ Sheet '{sheet_name}' not found in {file_name}. Available sheets: {', '.join(available_sheets)}")
                                            continue
                                    
                                    if df.empty:
                                        continue
                                    
                                    filtered = df.copy()
                                    selection_made = False
                                    
                                    # Apply filters from template
                                    for col in [df.columns[0], "PLANT", "APP", "NO"]:
                                        if col in row and pd.notna(row[col]) and str(row[col]).strip() != "":
                                            if col == "NO" and "," in str(row[col]):
                                                no_values = [v.strip() for v in str(row[col]).split(",")]
                                                filtered = filtered[filtered["NO"].astype(str).isin(no_values)]
                                            else:
                                                filtered = filtered[filtered[col].astype(str) == str(row[col])]
                                            selection_made = True
                                    
                                    if selection_made and not filtered.empty:
                                        filtered.insert(0, "User_ID", row["User_ID"])
                                        filtered.insert(1, "Actions", row.get("Actions", ""))
                                        filtered["Source_File"] = file_name
                                        filtered["Source_Sheet"] = sheet_name
                                        all_filtered.append(filtered)
                                except Exception as e:
                                    st.error(f"❌ Error processing {file_name} - {sheet_name}: {str(e)}")
                                    # Show available sheets for debugging
                                    try:
                                        temp_xls = pd.ExcelFile(source_file)
                                        st.info(f"📋 Available sheets in {file_name}: {', '.join(temp_xls.sheet_names)}")
                                    except:
                                        pass
                                    continue
                        
                        progress_bar.progress((sheet_idx + 1) / total_sheets)
                    
                    if all_filtered:
                        consolidated = pd.concat(all_filtered, ignore_index=True)
                        # Apply business rules
                        consolidated = consolidated[consolidated["Source_File"] != "PLANT ALL.xlsx"]
                        consolidated = consolidated.dropna(axis=1, how="all")
                        
                        # Ensure Total column is preserved in final results (if it exists in source data)
                        # The Total column will be included automatically from the source data
                        
                        st.session_state.consolidated_data = consolidated
                        st.success(f"✅ Successfully processed {len(consolidated)} records for {len(consolidated['User_ID'].unique())} users!")
                    else:
                        st.warning("⚠️ No valid data found in the template file.")
                
                except Exception as e:
                    st.error(f"Error processing mass upload: {str(e)}")
    
    st.markdown('</div>', unsafe_allow_html=True)

# Results Section
if st.session_state.consolidated_data is not None:
    st.markdown('<div class="results-container">', unsafe_allow_html=True)
    st.markdown('<div class="table-header">📊 Consolidated Results</div>', unsafe_allow_html=True)
    
    df = st.session_state.consolidated_data
    
    # Display summary metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Records", len(df))
    with col2:
        st.metric("Unique Users", df['User_ID'].nunique())
    with col3:
        st.metric("Source Files", df['Source_File'].nunique())
    with col4:
        st.metric("Sheets Processed", df['Source_Sheet'].nunique())
    
    # Display the consolidated data
    st.dataframe(df, use_container_width=True, height=400)
    
    # Download consolidated results
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Consolidated")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.download_button(
            label="📥 Download Consolidated Excel",
            data=output.getvalue(),
            file_name="consolidated_roles.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    
    st.markdown('</div>', unsafe_allow_html=True)

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #444; padding: 20px;'>"
    "Built with ❤️ using Streamlit | Excel Role Consolidator v2.0"
    "</div>",
    unsafe_allow_html=True
)
